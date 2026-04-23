#!/usr/bin/env python3
"""Extract tasks from the Orange Group work schedule Excel file into normalized JSON."""
import json
import openpyxl
from pathlib import Path
from datetime import datetime

SRC = Path(__file__).parent.parent / "source" / "schedule_original.xlsx"
OUT = Path(__file__).parent.parent / "data" / "tasks.json"


def parse_ru_date(s):
    if not s or not isinstance(s, str):
        return None
    parts = s.strip().split()
    for p in parts:
        try:
            return datetime.strptime(p, "%d.%m.%y").date().isoformat()
        except ValueError:
            continue
    return None


def parse_duration(v):
    if v is None or v == "-":
        return None
    if isinstance(v, int):
        return v
    s = str(v).strip()
    digits = "".join(ch for ch in s if ch.isdigit())
    return int(digits) if digits else None


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb.active

    tasks = []
    current_parent = None

    for r in range(3, 80):
        ident = ws.cell(row=r, column=1).value
        name = ws.cell(row=r, column=2).value
        dur = ws.cell(row=r, column=3).value
        start = ws.cell(row=r, column=4).value
        finish = ws.cell(row=r, column=5).value

        if not name:
            continue
        name = str(name).strip()

        if isinstance(start, datetime):
            start_iso = start.date().isoformat()
        else:
            start_iso = parse_ru_date(start)

        if isinstance(finish, datetime):
            finish_iso = finish.date().isoformat()
        else:
            finish_iso = parse_ru_date(finish)

        duration = parse_duration(dur)

        task = {
            "excel_row": r,
            "external_id": ident if isinstance(ident, int) else None,
            "name": name,
            "duration_days": duration,
            "start": start_iso,
            "finish": finish_iso,
            "parent": None,
        }

        if ident is not None:
            current_parent = name
        else:
            task["parent"] = current_parent

        tasks.append(task)

    holidays = [
        {"start": "2026-03-20", "end": "2026-03-22", "reason": "Праздничные дни"},
        {"start": "2026-05-26", "end": "2026-05-29", "reason": "Праздничные дни"},
        {"start": "2026-06-16", "end": "2026-06-16", "reason": "Праздничный день"},
    ]

    OUT.parent.mkdir(exist_ok=True)
    with OUT.open("w", encoding="utf-8") as f:
        json.dump({"tasks": tasks, "holidays": holidays}, f, ensure_ascii=False, indent=2)

    print(f"Wrote {len(tasks)} tasks and {len(holidays)} holiday ranges to {OUT}")
    print("\nParents (top-level tasks):")
    for t in tasks:
        if t["external_id"] is not None:
            print(f"  {t['external_id']:2}. {t['name']:50} {t['start']} → {t['finish']} ({t['duration_days']} дн)")


if __name__ == "__main__":
    main()
